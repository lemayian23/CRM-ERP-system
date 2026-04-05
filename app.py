from flask import Flask, render_template, request, session, redirect, url_for, send_file
from dotenv import load_dotenv
import os
import csv
import json
import tempfile
# import pdfkit
from io import BytesIO, StringIO
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from config.database import get_db

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')

# ============ HELPER FUNCTIONS ============

def get_dashboard_metrics():
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    metrics = {
        'unassigned_jc': 0, 'running_jc': 0, 'past_planned_jc': 0,
        'pending_payment_count': 0, 'pending_payment_total': 0,
        'customers': {'Individual': 0, 'Company': 0, 'Reseller': 0},
        'running_projects': 0, 'items_count': 0, 'products_count': 0,
        'items_value': 0, 'products_value': 0
    }
    
    try:
        cursor.execute("SELECT COUNT(*) as count FROM tbl_service_jc WHERE jc_assigned_to IS NULL OR jc_assigned_to = ''")
        row = cursor.fetchone()
        if row: metrics['unassigned_jc'] = row['count']
    except: pass
    
    try:
        cursor.execute("SELECT COUNT(*) as count FROM tbl_service_jc WHERE jc_assigned_to IS NOT NULL AND jc_assigned_to != '' AND (jc_closed IS NULL OR jc_closed != 'Closed') AND (paid IS NULL OR paid != 'Paid')")
        row = cursor.fetchone()
        if row: metrics['running_jc'] = row['count']
    except: pass
    
    try:
        cursor.execute("SELECT COUNT(*) as count FROM tbl_service_jc WHERE jc_assigned_to IS NOT NULL AND jc_assigned_to != '' AND (jc_closed IS NULL OR jc_closed != 'Closed') AND (paid IS NULL OR paid != 'Paid') AND proposed_work_date < %s", (date.today(),))
        row = cursor.fetchone()
        if row: metrics['past_planned_jc'] = row['count']
    except: pass
    
    try:
        cursor.execute("SELECT COUNT(*) as count, COALESCE(SUM(amount), 0) as total FROM tbl_service_jc WHERE jc_closed = 'Closed' AND (paid IS NULL OR paid != 'Paid')")
        row = cursor.fetchone()
        if row:
            metrics['pending_payment_count'] = row['count']
            metrics['pending_payment_total'] = float(row['total']) if row['total'] else 0
    except: pass
    
    try:
        cursor.execute("SELECT customer_type, COUNT(*) as count FROM tbl_customer GROUP BY customer_type")
        rows = cursor.fetchall()
        for row in rows:
            if row['customer_type'] in metrics['customers']:
                metrics['customers'][row['customer_type']] = row['count']
    except: pass
    
    try:
        cursor.execute("SELECT COUNT(*) as count FROM tbl_project WHERE status IN ('planning', 'running')")
        row = cursor.fetchone()
        if row: metrics['running_projects'] = row['count']
    except: pass
    
    try:
        cursor.execute("SELECT COUNT(*) as count FROM tbl_item")
        row = cursor.fetchone()
        if row: metrics['items_count'] = row['count']
    except: pass
    
    try:
        cursor.execute("SELECT COUNT(*) as count FROM tbl_product")
        row = cursor.fetchone()
        if row: metrics['products_count'] = row['count']
    except: pass
    
    try:
        cursor.execute("SELECT COALESCE(SUM(quantity * purchase_price), 0) as total FROM tbl_item")
        row = cursor.fetchone()
        if row: metrics['items_value'] = float(row['total']) if row['total'] else 0
    except: pass
    
    try:
        cursor.execute("SELECT COALESCE(SUM(quantity * purchase_price), 0) as total FROM tbl_product")
        row = cursor.fetchone()
        if row: metrics['products_value'] = float(row['total']) if row['total'] else 0
    except: pass
    
    cursor.close()
    conn.close()
    return metrics

def log_audit(action, table_name, record_id, old_values=None, new_values=None):
    if 'user_id' not in session:
        return
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO tbl_audit_log (user_id, username, action, table_name, record_id, old_values, new_values, ip_address, user_agent)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        session['user_id'], session['username'], action, table_name, record_id,
        json.dumps(old_values) if old_values else None,
        json.dumps(new_values) if new_values else None,
        request.remote_addr, request.headers.get('User-Agent', '')[:255]
    ))
    conn.commit()
    cursor.close()
    conn.close()

# ============ AUTHENTICATION ============

@app.route('/')
def home():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET'])
def login():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login_post():
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_admin WHERE username=%s AND password=%s",
                   (request.form['username'], request.form['password']))
    user = cursor.fetchone()
    cursor.close()
    conn.close()
    if user:
        session['user_id'] = user['user_id']
        session['user'] = user['user_type']
        session['username'] = user['username']
        return redirect(url_for('dashboard'))
    else:
        return render_template('login.html', error='Invalid username or password')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ============ DASHBOARD ============

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    metrics = get_dashboard_metrics()
    return render_template('dashboard.html', username=session['username'], role=session['user'],
                         metrics=metrics, active_page='dashboard')

# ============ JOB CARDS ============

@app.route('/job-cards')
def job_cards():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_service_jc ORDER BY id DESC")
    job_cards = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('job_cards.html', username=session['username'], role=session['user'],
                         job_cards=job_cards, active_page='job_cards')

@app.route('/job-cards/create', methods=['GET'])
def create_job_card_form():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('create_job_card.html', username=session['username'], role=session['user'],
                         active_page='job_cards')

@app.route('/job-cards/create', methods=['POST'])
def create_job_card_post():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO tbl_service_jc (jc_type, customer_type, customer_name, customer_id, jc_create_date, amount, work_statement)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (request.form['jc_type'], request.form['customer_type'], request.form['customer_name'],
          request.form['customer_id'] or None, request.form['jc_create_date'], request.form['amount'], request.form['work_statement']))
    jc_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_service_jc', jc_id, None, {'customer_name': request.form['customer_name'], 'amount': request.form['amount']})
    cursor.close()
    conn.close()
    return redirect(url_for('job_cards'))

@app.route('/job-cards/<int:jc_id>')
def view_job_card(jc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_service_jc WHERE id = %s", (jc_id,))
    job_card = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('view_job_card.html', username=session['username'], role=session['user'],
                         job_card=job_card, active_page='job_cards')

@app.route('/job-cards/<int:jc_id>/edit', methods=['GET'])
def edit_job_card_form(jc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_service_jc WHERE id = %s", (jc_id,))
    job_card = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('edit_job_card.html', username=session['username'], role=session['user'],
                         job_card=job_card, active_page='job_cards')

@app.route('/job-cards/<int:jc_id>/edit', methods=['POST'])
def edit_job_card_post(jc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_service_jc SET customer_name = %s, amount = %s, work_statement = %s WHERE id = %s",
                   (request.form['customer_name'], request.form['amount'], request.form['work_statement'], jc_id))
    conn.commit()
    log_audit('UPDATE', 'tbl_service_jc', jc_id, None, {'customer_name': request.form['customer_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('job_cards'))

@app.route('/job-cards/<int:jc_id>/assign', methods=['GET'])
def assign_technician_form(jc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_service_jc WHERE id = %s", (jc_id,))
    job_card = cursor.fetchone()
    cursor.execute("SELECT * FROM tbl_technician WHERE status = 'active'")
    technicians = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('assign_technician.html', username=session['username'], role=session['user'],
                         job_card=job_card, technicians=technicians, active_page='job_cards')

@app.route('/job-cards/<int:jc_id>/assign', methods=['POST'])
def assign_technician_post(jc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_service_jc SET jc_assigned_to = %s, proposed_work_date = %s, time_slot = %s WHERE id = %s",
                   (request.form['technician'], request.form['proposed_work_date'], request.form['time_slot'], jc_id))
    conn.commit()
    log_audit('UPDATE', 'tbl_service_jc', jc_id, None, {'jc_assigned_to': request.form['technician']})
    cursor.close()
    conn.close()
    return redirect(url_for('job_cards'))

@app.route('/job-cards/<int:jc_id>/close', methods=['GET'])
def close_job_card_form(jc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_service_jc WHERE id = %s", (jc_id,))
    job_card = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('close_job_card.html', username=session['username'], role=session['user'],
                         job_card=job_card, active_page='job_cards')

@app.route('/job-cards/<int:jc_id>/close', methods=['POST'])
def close_job_card_post(jc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    # Get job card items first
    cursor.execute("SELECT * FROM tbl_service_jc_item WHERE service_jc_id = %s", (jc_id,))
    items = cursor.fetchall()
    
    # Deduct stock for each item
    for item in items:
        if item['item_id']:
            # Deduct from tbl_item
            cursor.execute("UPDATE tbl_item SET quantity = quantity - %s WHERE item_id = %s",
                          (item['item_quantity'], item['item_id']))
            # Record stock movement
            cursor.execute("""
                INSERT INTO tbl_stock_movement_item (item_id, item_name, movement_type, quantity, 
                previous_quantity, new_quantity, reference_type, reference_id, notes, created_by)
                SELECT item_id, item_name, 'OUT', %s, quantity, quantity - %s, 'JC_CLOSE', %s, %s, %s
                FROM tbl_item WHERE item_id = %s
            """, (item['item_quantity'], item['item_quantity'], jc_id, 
                  f"Job Card #{jc_id} closure", session['username'], item['item_id']))
        
        if item['product_id']:
            # Deduct from tbl_product
            cursor.execute("UPDATE tbl_product SET quantity = quantity - %s WHERE product_id = %s",
                          (item['product_quantity'], item['product_id']))
            # Record stock movement
            cursor.execute("""
                INSERT INTO tbl_stock_movement_product (product_id, product_name, movement_type, quantity,
                previous_quantity, new_quantity, reference_type, reference_id, notes, created_by)
                SELECT product_id, product_name, 'OUT', %s, quantity, quantity - %s, 'JC_CLOSE', %s, %s, %s
                FROM tbl_product WHERE product_id = %s
            """, (item['product_quantity'], item['product_quantity'], jc_id,
                  f"Job Card #{jc_id} closure", session['username'], item['product_id']))
    
    # Update job card as closed
    cursor.execute("""
        UPDATE tbl_service_jc 
        SET jc_closed = 'Closed', job_finding = %s, work_done_date = %s, hours = %s
        WHERE id = %s
    """, (request.form['job_finding'], request.form['work_done_date'], request.form['hours'] or 0, jc_id))
    
    conn.commit()
    log_audit('UPDATE', 'tbl_service_jc', jc_id, None, {'jc_closed': 'Closed', 'stock_deducted': 'Yes'})
    cursor.close()
    conn.close()
    
    return redirect(url_for('job_cards'))

@app.route('/job-cards/<int:jc_id>/payment', methods=['GET'])
def payment_job_card_form(jc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_service_jc WHERE id = %s", (jc_id,))
    job_card = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('payment_job_card.html', username=session['username'], role=session['user'],
                         job_card=job_card, active_page='job_cards')

@app.route('/job-cards/<int:jc_id>/payment', methods=['POST'])
def payment_job_card_post(jc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_service_jc SET paid = 'Paid', payment_type = %s, payment_code = %s, total_paid_amount = %s, payment_date = %s WHERE id = %s",
                   (request.form['payment_type'], request.form['payment_code'], request.form['total_paid_amount'], request.form['payment_date'], jc_id))
    conn.commit()
    log_audit('UPDATE', 'tbl_service_jc', jc_id, None, {'paid': 'Paid', 'payment_type': request.form['payment_type']})
    cursor.close()
    conn.close()
    return redirect(url_for('job_cards'))

# ============ CUSTOMERS ============

@app.route('/customers')
def customers():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_customer ORDER BY customer_id DESC")
    customers = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('customers.html', username=session['username'], role=session['user'],
                         customers=customers, active_page='customers')

@app.route('/customers/add', methods=['POST'])
def add_customer():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_customer (customer_name, customer_type, email, phone) VALUES (%s, %s, %s, %s)",
                   (request.form['customer_name'], request.form['customer_type'], request.form['email'], request.form['phone']))
    customer_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_customer', customer_id, None, {'customer_name': request.form['customer_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('customers'))

@app.route('/customers/edit', methods=['POST'])
def edit_customer():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_customer SET customer_name = %s, customer_type = %s, email = %s, phone = %s WHERE customer_id = %s",
                   (request.form['customer_name'], request.form['customer_type'], request.form['email'], request.form['phone'], request.form['customer_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_customer', request.form['customer_id'], None, {'customer_name': request.form['customer_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('customers'))

@app.route('/customers/delete/<int:customer_id>')
def delete_customer(customer_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_customer WHERE customer_id = %s", (customer_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_customer', customer_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('customers'))

# ============ INVENTORY ============

@app.route('/inventory')
def inventory():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_item ORDER BY item_id DESC")
    items = cursor.fetchall()
    cursor.execute("SELECT * FROM tbl_product ORDER BY product_id DESC")
    products = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('inventory.html', username=session['username'], role=session['user'],
                         items=items, products=products, active_page='inventory')

@app.route('/inventory/item/add', methods=['POST'])
def add_item():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_item (item_name, quantity, purchase_price) VALUES (%s, %s, %s)",
                   (request.form['item_name'], request.form['quantity'], request.form['purchase_price']))
    item_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_item', item_id, None, {'item_name': request.form['item_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('inventory'))

@app.route('/inventory/item/edit', methods=['POST'])
def edit_item():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_item SET item_name = %s, quantity = %s, purchase_price = %s WHERE item_id = %s",
                   (request.form['item_name'], request.form['quantity'], request.form['purchase_price'], request.form['item_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_item', request.form['item_id'], None, {'item_name': request.form['item_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('inventory'))

@app.route('/inventory/item/delete/<int:item_id>')
def delete_item(item_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_item WHERE item_id = %s", (item_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_item', item_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('inventory'))

@app.route('/inventory/product/add', methods=['POST'])
def add_product():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_product (product_name, quantity, purchase_price) VALUES (%s, %s, %s)",
                   (request.form['product_name'], request.form['quantity'], request.form['purchase_price']))
    product_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_product', product_id, None, {'product_name': request.form['product_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('inventory'))

@app.route('/inventory/product/edit', methods=['POST'])
def edit_product():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_product SET product_name = %s, quantity = %s, purchase_price = %s WHERE product_id = %s",
                   (request.form['product_name'], request.form['quantity'], request.form['purchase_price'], request.form['product_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_product', request.form['product_id'], None, {'product_name': request.form['product_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('inventory'))

@app.route('/inventory/product/delete/<int:product_id>')
def delete_product(product_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_product WHERE product_id = %s", (product_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_product', product_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('inventory'))

# ============ TECHNICIANS ============

@app.route('/technicians')
def technicians():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_technician ORDER BY tech_id DESC")
    technicians = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('technicians.html', username=session['username'], role=session['user'],
                         technicians=technicians, active_page='technicians')

@app.route('/technicians/add', methods=['POST'])
def add_technician():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_technician (tech_name, email, phone, status) VALUES (%s, %s, %s, %s)",
                   (request.form['tech_name'], request.form['email'], request.form['phone'], request.form['status']))
    tech_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_technician', tech_id, None, {'tech_name': request.form['tech_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('technicians'))

@app.route('/technicians/edit', methods=['POST'])
def edit_technician():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_technician SET tech_name = %s, email = %s, phone = %s, status = %s WHERE tech_id = %s",
                   (request.form['tech_name'], request.form['email'], request.form['phone'], request.form['status'], request.form['tech_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_technician', request.form['tech_id'], None, {'tech_name': request.form['tech_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('technicians'))

@app.route('/technicians/delete/<int:tech_id>')
def delete_technician(tech_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_technician WHERE tech_id = %s", (tech_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_technician', tech_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('technicians'))

# ============ REPORTS ============

@app.route('/reports')
def reports():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('reports.html', username=session['username'], role=session['user'], active_page='reports')

@app.route('/reports/financial')
def financial_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    period = request.args.get('period', 'all')
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    date_filter = ""
    if period == 'week': date_filter = "AND created_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)"
    elif period == 'month': date_filter = "AND created_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)"
    elif period == 'year': date_filter = "AND created_at >= DATE_SUB(NOW(), INTERVAL 365 DAY)"
    cursor.execute(f"SELECT * FROM tbl_service_jc WHERE 1=1 {date_filter} ORDER BY created_at DESC")
    payments = cursor.fetchall()
    cursor.execute(f"""
        SELECT COALESCE(SUM(CASE WHEN paid = 'Paid' THEN total_paid_amount ELSE amount END), 0) as total_revenue,
               COALESCE(SUM(CASE WHEN paid != 'Paid' AND jc_closed = 'Closed' THEN amount ELSE 0 END), 0) as pending_amount,
               COUNT(CASE WHEN paid = 'Paid' THEN 1 END) as paid_count, COUNT(*) as total_jobs,
               COUNT(CASE WHEN paid != 'Paid' AND jc_closed = 'Closed' THEN 1 END) as pending_count
        FROM tbl_service_jc WHERE 1=1 {date_filter}
    """)
    summary = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('financial_report.html', username=session['username'], role=session['user'],
                         payments=payments, summary=summary, active_page='reports')

@app.route('/reports/job-cards')
def job_card_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_service_jc ORDER BY created_at DESC")
    job_cards = cursor.fetchall()
    cursor.execute("""
        SELECT COUNT(*) as total, SUM(CASE WHEN jc_closed = 'Closed' THEN 1 ELSE 0 END) as closed,
               SUM(CASE WHEN jc_closed IS NULL OR jc_closed != 'Closed' THEN 1 ELSE 0 END) as open,
               SUM(CASE WHEN paid = 'Paid' THEN 1 ELSE 0 END) as paid,
               SUM(CASE WHEN paid != 'Paid' THEN 1 ELSE 0 END) as unpaid
        FROM tbl_service_jc
    """)
    summary = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('job_card_report.html', username=session['username'], role=session['user'],
                         job_cards=job_cards, summary=summary, active_page='reports')

@app.route('/reports/stock')
def stock_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_item ORDER BY item_name")
    items = cursor.fetchall()
    cursor.execute("SELECT * FROM tbl_product ORDER BY product_name")
    products = cursor.fetchall()
    cursor.execute("SELECT COALESCE(SUM(quantity * purchase_price), 0) as total FROM tbl_item")
    items_total = cursor.fetchone()
    cursor.execute("SELECT COALESCE(SUM(quantity * purchase_price), 0) as total FROM tbl_product")
    products_total = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('stock_report.html', username=session['username'], role=session['user'],
                         items=items, products=products, items_total=items_total['total'],
                         products_total=products_total['total'], active_page='reports')

@app.route('/reports/customers')
def customer_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT c.*, COUNT(jc.id) as total_jobs, SUM(CASE WHEN jc.paid = 'Paid' THEN jc.amount ELSE 0 END) as total_spent
        FROM tbl_customer c LEFT JOIN tbl_service_jc jc ON c.customer_name = jc.customer_name
        GROUP BY c.customer_id ORDER BY c.customer_name
    """)
    customers = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('customer_report.html', username=session['username'], role=session['user'],
                         customers=customers, active_page='reports')

@app.route('/reports/technicians')
def technician_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT t.*, COUNT(jc.id) as total_jobs, SUM(CASE WHEN jc.paid = 'Paid' THEN jc.amount ELSE 0 END) as revenue,
               SUM(CASE WHEN jc.jc_closed = 'Closed' THEN 1 ELSE 0 END) as completed_jobs
        FROM tbl_technician t LEFT JOIN tbl_service_jc jc ON t.tech_name = jc.jc_assigned_to
        GROUP BY t.tech_id ORDER BY t.tech_name
    """)
    technicians = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('technician_report.html', username=session['username'], role=session['user'],
                         technicians=technicians, active_page='reports')

# ============ MPESA NUMBERS ============

@app.route('/mpesa')
def mpesa_numbers():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_mpesa_number ORDER BY id DESC")
    mpesa_numbers = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('mpesa.html', username=session['username'], role=session['user'],
                         mpesa_numbers=mpesa_numbers, active_page='mpesa')

@app.route('/mpesa/add', methods=['POST'])
def add_mpesa():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_mpesa_number (name, number) VALUES (%s, %s)",
                   (request.form['name'], request.form['number']))
    mpesa_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_mpesa_number', mpesa_id, None, {'name': request.form['name']})
    cursor.close()
    conn.close()
    return redirect(url_for('mpesa_numbers'))

@app.route('/mpesa/edit', methods=['POST'])
def edit_mpesa():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_mpesa_number SET name = %s, number = %s WHERE id = %s",
                   (request.form['name'], request.form['number'], request.form['id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_mpesa_number', request.form['id'], None, {'name': request.form['name']})
    cursor.close()
    conn.close()
    return redirect(url_for('mpesa_numbers'))

@app.route('/mpesa/delete/<int:mpesa_id>')
def delete_mpesa(mpesa_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_mpesa_number WHERE id = %s", (mpesa_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_mpesa_number', mpesa_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('mpesa_numbers'))

# ============ USERS ============

@app.route('/users')
def users():
    if 'user_id' not in session or session.get('user') != 'admin':
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_admin ORDER BY user_id DESC")
    users = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('users.html', username=session['username'], role=session['user'],
                         users=users, active_page='users')

@app.route('/users/add', methods=['POST'])
def add_user():
    if 'user_id' not in session or session.get('user') != 'admin':
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_admin (username, password, user, email) VALUES (%s, %s, %s, %s)",
                   (request.form['username'], request.form['password'], request.form['role'], request.form['email']))
    user_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_admin', user_id, None, {'username': request.form['username']})
    cursor.close()
    conn.close()
    return redirect(url_for('users'))

@app.route('/users/edit', methods=['POST'])
def edit_user():
    if 'user_id' not in session or session.get('user') != 'admin':
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    if request.form['password']:
        cursor.execute("UPDATE tbl_admin SET username = %s, password = %s, user = %s, email = %s WHERE user_id = %s",
                       (request.form['username'], request.form['password'], request.form['role'], request.form['email'], request.form['user_id']))
    else:
        cursor.execute("UPDATE tbl_admin SET username = %s, user = %s, email = %s WHERE user_id = %s",
                       (request.form['username'], request.form['role'], request.form['email'], request.form['user_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_admin', request.form['user_id'], None, {'username': request.form['username']})
    cursor.close()
    conn.close()
    return redirect(url_for('users'))

@app.route('/users/delete/<int:user_id>')
def delete_user(user_id):
    if 'user_id' not in session or session.get('user') != 'admin' or user_id == session['user_id']:
        return redirect(url_for('users'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_admin WHERE user_id = %s", (user_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_admin', user_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('users'))

# ============ PROJECTS ============

@app.route('/projects')
def projects():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_project ORDER BY project_id DESC")
    projects = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('projects.html', username=session['username'], role=session['user'],
                         projects=projects, active_page='projects')

@app.route('/projects/add', methods=['POST'])
def add_project():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO tbl_project (project_name, description, start_date, end_date, status, budget, customer_id, assigned_to)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """, (request.form['project_name'], request.form['project_description'], request.form['start_date'] or None,
          request.form['end_date'] or None, request.form['status'], request.form['budget'] or 0,
          request.form['customer_id'] or None, request.form['assigned_to'] or None))
    project_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_project', project_id, None, {'project_name': request.form['project_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('projects'))

@app.route('/projects/edit', methods=['POST'])
def edit_project():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE tbl_project SET project_name = %s, description = %s, start_date = %s, end_date = %s,
        status = %s, budget = %s, customer_id = %s, assigned_to = %s WHERE project_id = %s
    """, (request.form['project_name'], request.form['project_description'], request.form['start_date'] or None,
          request.form['end_date'] or None, request.form['status'], request.form['budget'] or 0,
          request.form['customer_id'] or None, request.form['assigned_to'] or None, request.form['project_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_project', request.form['project_id'], None, {'project_name': request.form['project_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('projects'))

@app.route('/projects/delete/<int:project_id>')
def delete_project(project_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_project WHERE project_id = %s", (project_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_project', project_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('projects'))

@app.route('/projects/view/<int:project_id>')
def view_project(project_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_project WHERE project_id = %s", (project_id,))
    project = cursor.fetchone()
    cursor.execute("SELECT * FROM tbl_service_jc WHERE customer_id = %s OR customer_name LIKE %s ORDER BY created_at DESC LIMIT 10",
                   (project.get('customer_id'), f"%{project.get('project_name')}%"))
    job_cards = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('view_project.html', username=session['username'], role=session['user'],
                         project=project, job_cards=job_cards, active_page='projects')

# ============ STOCK MOVEMENT ============

@app.route('/stock-movement')
def stock_movement():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_stock_movement_item ORDER BY created_at DESC LIMIT 100")
    item_movements = cursor.fetchall()
    cursor.execute("SELECT * FROM tbl_stock_movement_product ORDER BY created_at DESC LIMIT 100")
    product_movements = cursor.fetchall()
    cursor.execute("SELECT * FROM tbl_item ORDER BY item_name")
    items = cursor.fetchall()
    cursor.execute("SELECT * FROM tbl_product ORDER BY product_name")
    products = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('stock_movement.html', username=session['username'], role=session['user'],
                         item_movements=item_movements, product_movements=product_movements,
                         items=items, products=products, active_page='stock_movement')

@app.route('/stock-movement/adjust-item', methods=['POST'])
def adjust_item_stock():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    item_id = request.form['item_id']
    new_quantity = int(request.form['new_quantity'])
    reason = request.form['reason']
    cursor.execute("SELECT item_name, quantity FROM tbl_item WHERE item_id = %s", (item_id,))
    item = cursor.fetchone()
    if item:
        old_quantity = item['quantity']
        cursor.execute("UPDATE tbl_item SET quantity = %s WHERE item_id = %s", (new_quantity, item_id))
        cursor.execute("INSERT INTO tbl_stock_movement_item (item_id, item_name, movement_type, quantity, previous_quantity, new_quantity, reference_type, notes, created_by) VALUES (%s, %s, 'ADJUSTMENT', %s, %s, %s, 'CORRECTION', %s, %s)",
                       (item_id, item['item_name'], abs(new_quantity - old_quantity), old_quantity, new_quantity, reason, session['username']))
        conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('stock_movement'))

@app.route('/stock-movement/adjust-product', methods=['POST'])
def adjust_product_stock():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    product_id = request.form['product_id']
    new_quantity = int(request.form['new_quantity'])
    reason = request.form['reason']
    cursor.execute("SELECT product_name, quantity FROM tbl_product WHERE product_id = %s", (product_id,))
    product = cursor.fetchone()
    if product:
        old_quantity = product['quantity']
        cursor.execute("UPDATE tbl_product SET quantity = %s WHERE product_id = %s", (new_quantity, product_id))
        cursor.execute("INSERT INTO tbl_stock_movement_product (product_id, product_name, movement_type, quantity, previous_quantity, new_quantity, reference_type, notes, created_by) VALUES (%s, %s, 'ADJUSTMENT', %s, %s, %s, 'CORRECTION', %s, %s)",
                       (product_id, product['product_name'], abs(new_quantity - old_quantity), old_quantity, new_quantity, reason, session['username']))
        conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('stock_movement'))

# ============ CATEGORIES ============

@app.route('/categories')
def categories():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_category ORDER BY category_id DESC")
    categories = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('categories.html', username=session['username'], role=session['user'],
                         categories=categories, active_page='categories')

@app.route('/categories/add', methods=['POST'])
def add_category():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_category (category_name, description) VALUES (%s, %s)",
                   (request.form['category_name'], request.form['description']))
    category_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_category', category_id, None, {'category_name': request.form['category_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('categories'))

@app.route('/categories/edit', methods=['POST'])
def edit_category():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_category SET category_name = %s, description = %s WHERE category_id = %s",
                   (request.form['category_name'], request.form['description'], request.form['category_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_category', request.form['category_id'], None, {'category_name': request.form['category_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('categories'))

@app.route('/categories/delete/<int:category_id>')
def delete_category(category_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_category WHERE category_id = %s", (category_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_category', category_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('categories'))

# ============ BRANDS ============

@app.route('/brands')
def brands():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_brand ORDER BY brand_id DESC")
    brands = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('brands.html', username=session['username'], role=session['user'],
                         brands=brands, active_page='brands')

@app.route('/brands/add', methods=['POST'])
def add_brand():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_brand (brand_name, description) VALUES (%s, %s)",
                   (request.form['brand_name'], request.form['description']))
    brand_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_brand', brand_id, None, {'brand_name': request.form['brand_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('brands'))

@app.route('/brands/edit', methods=['POST'])
def edit_brand():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_brand SET brand_name = %s, description = %s WHERE brand_id = %s",
                   (request.form['brand_name'], request.form['description'], request.form['brand_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_brand', request.form['brand_id'], None, {'brand_name': request.form['brand_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('brands'))

@app.route('/brands/delete/<int:brand_id>')
def delete_brand(brand_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_brand WHERE brand_id = %s", (brand_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_brand', brand_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('brands'))

# ============ SUBCATEGORIES ============

@app.route('/subcategories')
def subcategories():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT s.*, c.category_name FROM tbl_subcategory s LEFT JOIN tbl_category c ON s.category_id = c.category_id ORDER BY s.subcategory_id DESC")
    subcategories = cursor.fetchall()
    cursor.execute("SELECT category_id, category_name FROM tbl_category")
    categories = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('subcategories.html', username=session['username'], role=session['user'],
                         subcategories=subcategories, categories=categories, active_page='subcategories')

@app.route('/subcategories/add', methods=['POST'])
def add_subcategory():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_subcategory (subcategory_name, category_id, description) VALUES (%s, %s, %s)",
                   (request.form['subcategory_name'], request.form['category_id'] or None, request.form['description']))
    sub_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_subcategory', sub_id, None, {'subcategory_name': request.form['subcategory_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('subcategories'))

@app.route('/subcategories/edit', methods=['POST'])
def edit_subcategory():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_subcategory SET subcategory_name = %s, category_id = %s, description = %s WHERE subcategory_id = %s",
                   (request.form['subcategory_name'], request.form['category_id'] or None, request.form['description'], request.form['subcategory_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_subcategory', request.form['subcategory_id'], None, {'subcategory_name': request.form['subcategory_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('subcategories'))

@app.route('/subcategories/delete/<int:subcategory_id>')
def delete_subcategory(subcategory_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_subcategory WHERE subcategory_id = %s", (subcategory_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_subcategory', subcategory_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('subcategories'))

# ============ BUSINESS UNITS ============

@app.route('/business-units')
def business_units():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_business ORDER BY business_id DESC")
    business_units = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('business_units.html', username=session['username'], role=session['user'],
                         business_units=business_units, active_page='business_units')

@app.route('/business-units/add', methods=['POST'])
def add_business_unit():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_business (business_name, description) VALUES (%s, %s)",
                   (request.form['business_name'], request.form['description']))
    business_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_business', business_id, None, {'business_name': request.form['business_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('business_units'))

@app.route('/business-units/edit', methods=['POST'])
def edit_business_unit():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_business SET business_name = %s, description = %s WHERE business_id = %s",
                   (request.form['business_name'], request.form['description'], request.form['business_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_business', request.form['business_id'], None, {'business_name': request.form['business_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('business_units'))

@app.route('/business-units/delete/<int:business_id>')
def delete_business_unit(business_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_business WHERE business_id = %s", (business_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_business', business_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('business_units'))

# ============ PRICE LISTS ============

@app.route('/price-lists')
def price_lists():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_item ORDER BY item_name")
    items = cursor.fetchall()
    cursor.execute("SELECT * FROM tbl_product ORDER BY product_name")
    products = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('price_lists.html', username=session['username'], role=session['user'],
                         items=items, products=products, active_page='price_lists')

@app.route('/price-lists/update-item', methods=['POST'])
def update_item_price():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_item SET purchase_price = %s WHERE item_id = %s",
                   (request.form['purchase_price'], request.form['item_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_item', request.form['item_id'], None, {'purchase_price': request.form['purchase_price']})
    cursor.close()
    conn.close()
    return redirect(url_for('price_lists'))

@app.route('/price-lists/update-product', methods=['POST'])
def update_product_price():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_product SET purchase_price = %s WHERE product_id = %s",
                   (request.form['purchase_price'], request.form['product_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_product', request.form['product_id'], None, {'purchase_price': request.form['purchase_price']})
    cursor.close()
    conn.close()
    return redirect(url_for('price_lists'))

# ============ RATES ============

@app.route('/rates')
def rates():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_rates ORDER BY rate_id DESC")
    rates = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('rates.html', username=session['username'], role=session['user'],
                         rates=rates, active_page='rates')

@app.route('/rates/add', methods=['POST'])
def add_rate():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_rates (rate_name, rate_type, amount, description, status, created_by) VALUES (%s, %s, %s, %s, %s, %s)",
                   (request.form['rate_name'], request.form['rate_type'], request.form['amount'], request.form['description'], request.form['status'], session['username']))
    rate_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_rates', rate_id, None, {'rate_name': request.form['rate_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('rates'))

@app.route('/rates/edit', methods=['POST'])
def edit_rate():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_rates SET rate_name = %s, rate_type = %s, amount = %s, description = %s, status = %s WHERE rate_id = %s",
                   (request.form['rate_name'], request.form['rate_type'], request.form['amount'], request.form['description'], request.form['status'], request.form['rate_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_rates', request.form['rate_id'], None, {'rate_name': request.form['rate_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('rates'))

@app.route('/rates/delete/<int:rate_id>')
def delete_rate(rate_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_rates WHERE rate_id = %s", (rate_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_rates', rate_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('rates'))

# ============ SERVICE CALLS ============

@app.route('/service-calls')
def service_calls():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_service_calls ORDER BY call_id DESC")
    calls = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('service_calls.html', username=session['username'], role=session['user'],
                         calls=calls, active_page='service_calls')

@app.route('/service-calls/add', methods=['POST'])
def add_service_call():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_service_calls (customer_name, customer_phone, call_date, issue_description, priority, status, created_by) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                   (request.form['customer_name'], request.form['customer_phone'], request.form['call_date'], request.form['issue_description'], request.form['priority'], 'pending', session['username']))
    call_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_service_calls', call_id, None, {'customer_name': request.form['customer_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('service_calls'))

@app.route('/service-calls/edit', methods=['POST'])
def edit_service_call():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_service_calls SET customer_name = %s, customer_phone = %s, issue_description = %s, priority = %s, status = %s, assigned_to = %s WHERE call_id = %s",
                   (request.form['customer_name'], request.form['customer_phone'], request.form['issue_description'], request.form['priority'], request.form['status'], request.form['assigned_to'], request.form['call_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_service_calls', request.form['call_id'], None, {'customer_name': request.form['customer_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('service_calls'))

@app.route('/service-calls/resolve', methods=['POST'])
def resolve_service_call():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_service_calls SET status = 'resolved', resolution_notes = %s, resolved_date = %s WHERE call_id = %s",
                   (request.form['resolution_notes'], request.form['resolved_date'], request.form['call_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_service_calls', request.form['call_id'], None, {'status': 'resolved'})
    cursor.close()
    conn.close()
    return redirect(url_for('service_calls'))

@app.route('/service-calls/delete/<int:call_id>')
def delete_service_call(call_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_service_calls WHERE call_id = %s", (call_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_service_calls', call_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('service_calls'))

# ============ AMC CONTRACTS ============

@app.route('/amc')
def amc_contracts():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_amc ORDER BY amc_id DESC")
    contracts = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('amc.html', username=session['username'], role=session['user'],
                         contracts=contracts, active_page='amc')

@app.route('/amc/add', methods=['POST'])
def add_amc():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    amc_number = f"AMC-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    cursor.execute("""
        INSERT INTO tbl_amc (amc_number, customer_id, customer_name, contract_start, contract_end, contract_value, payment_terms, service_level, status, notes, created_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (amc_number, request.form['customer_id'] or None, request.form['customer_name'], request.form['contract_start'],
          request.form['contract_end'], request.form['contract_value'], request.form['payment_terms'], request.form['service_level'],
          request.form['status'], request.form['notes'], session['username']))
    amc_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_amc', amc_id, None, {'amc_number': amc_number})
    cursor.close()
    conn.close()
    return redirect(url_for('amc_contracts'))

@app.route('/amc/edit', methods=['POST'])
def edit_amc():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE tbl_amc SET customer_id = %s, customer_name = %s, contract_start = %s, contract_end = %s,
        contract_value = %s, payment_terms = %s, service_level = %s, status = %s, notes = %s WHERE amc_id = %s
    """, (request.form['customer_id'] or None, request.form['customer_name'], request.form['contract_start'],
          request.form['contract_end'], request.form['contract_value'], request.form['payment_terms'],
          request.form['service_level'], request.form['status'], request.form['notes'], request.form['amc_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_amc', request.form['amc_id'], None, {'customer_name': request.form['customer_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('amc_contracts'))

@app.route('/amc/delete/<int:amc_id>')
def delete_amc(amc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_amc WHERE amc_id = %s", (amc_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_amc', amc_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('amc_contracts'))

@app.route('/amc/view/<int:amc_id>')
def view_amc(amc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_amc WHERE amc_id = %s", (amc_id,))
    contract = cursor.fetchone()
    cursor.execute("""
        SELECT aj.*, jc.jc_number, jc.amount, jc.jc_closed, jc.paid
        FROM tbl_amc_jc aj LEFT JOIN tbl_service_jc jc ON aj.jc_id = jc.id
        WHERE aj.amc_id = %s ORDER BY aj.visit_date DESC
    """, (amc_id,))
    amc_jobs = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('view_amc.html', username=session['username'], role=session['user'],
                         contract=contract, amc_jobs=amc_jobs, active_page='amc')

@app.route('/amc/add-job/<int:amc_id>', methods=['POST'])
def add_amc_job(amc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_service_jc (customer_name, jc_type, amount, work_statement, created_by) VALUES (%s, 'AMC', %s, %s, %s)",
                   (request.form['customer_name'], request.form['amount'] or 0, request.form['work_done'], session['username']))
    jc_id = cursor.lastrowid
    cursor.execute("INSERT INTO tbl_amc_jc (amc_id, jc_id, visit_date, technician_name, work_done, status) VALUES (%s, %s, %s, %s, %s, 'completed')",
                   (amc_id, jc_id, request.form['visit_date'], request.form['technician_name'], request.form['work_done']))
    conn.commit()
    log_audit('CREATE', 'tbl_amc_jc', jc_id, None, {'amc_id': amc_id})
    cursor.close()
    conn.close()
    return redirect(url_for('view_amc', amc_id=amc_id))

# ============ AREAS ============

@app.route('/areas')
def areas():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_area ORDER BY area_id DESC")
    areas = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('areas.html', username=session['username'], role=session['user'],
                         areas=areas, active_page='areas')

@app.route('/areas/add', methods=['POST'])
def add_area():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_area (area_name, zone, distance_km, travel_cost, status) VALUES (%s, %s, %s, %s, %s)",
                   (request.form['area_name'], request.form['zone'], request.form['distance_km'] or 0, request.form['travel_cost'] or 0, request.form['status']))
    area_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_area', area_id, None, {'area_name': request.form['area_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('areas'))

@app.route('/areas/edit', methods=['POST'])
def edit_area():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_area SET area_name = %s, zone = %s, distance_km = %s, travel_cost = %s, status = %s WHERE area_id = %s",
                   (request.form['area_name'], request.form['zone'], request.form['distance_km'] or 0, request.form['travel_cost'] or 0, request.form['status'], request.form['area_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_area', request.form['area_id'], None, {'area_name': request.form['area_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('areas'))

@app.route('/areas/delete/<int:area_id>')
def delete_area(area_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_area WHERE area_id = %s", (area_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_area', area_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('areas'))

# ============ COMMERCE ============

@app.route('/commerce')
def commerce():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT c.*, a.area_name FROM tbl_commerce c LEFT JOIN tbl_area a ON c.area_id = a.area_id ORDER BY c.commerce_id DESC")
    commerce = cursor.fetchall()
    cursor.execute("SELECT area_id, area_name FROM tbl_area WHERE status = 'active'")
    areas = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('commerce.html', username=session['username'], role=session['user'],
                         commerce=commerce, areas=areas, active_page='commerce')

@app.route('/commerce/add', methods=['POST'])
def add_commerce():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tbl_commerce (commerce_name, area_id, address, contact_person, phone, email, status) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                   (request.form['commerce_name'], request.form['area_id'] or None, request.form['address'], request.form['contact_person'], request.form['phone'], request.form['email'], request.form['status']))
    commerce_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_commerce', commerce_id, None, {'commerce_name': request.form['commerce_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('commerce'))

@app.route('/commerce/edit', methods=['POST'])
def edit_commerce():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_commerce SET commerce_name = %s, area_id = %s, address = %s, contact_person = %s, phone = %s, email = %s, status = %s WHERE commerce_id = %s",
                   (request.form['commerce_name'], request.form['area_id'] or None, request.form['address'], request.form['contact_person'], request.form['phone'], request.form['email'], request.form['status'], request.form['commerce_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_commerce', request.form['commerce_id'], None, {'commerce_name': request.form['commerce_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('commerce'))

@app.route('/commerce/delete/<int:commerce_id>')
def delete_commerce(commerce_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_commerce WHERE commerce_id = %s", (commerce_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_commerce', commerce_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('commerce'))

# ============ SALES AGENTS ============

@app.route('/sales-agents')
def sales_agents():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_sales_agents ORDER BY agent_id DESC")
    agents = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('sales_agents.html', username=session['username'], role=session['user'],
                         agents=agents, active_page='sales_agents')

@app.route('/sales-agents/add', methods=['POST'])
def add_sales_agent():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO tbl_sales_agents (agent_name, email, phone, commission_rate, target_amount, status, hire_date, notes, created_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (request.form['agent_name'], request.form['email'], request.form['phone'], request.form['commission_rate'] or 0,
          request.form['target_amount'] or 0, request.form['status'], request.form['hire_date'] or None,
          request.form['notes'], session['username']))
    agent_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_sales_agents', agent_id, None, {'agent_name': request.form['agent_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('sales_agents'))

@app.route('/sales-agents/edit', methods=['POST'])
def edit_sales_agent():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE tbl_sales_agents SET agent_name = %s, email = %s, phone = %s, commission_rate = %s,
        target_amount = %s, status = %s, hire_date = %s, notes = %s WHERE agent_id = %s
    """, (request.form['agent_name'], request.form['email'], request.form['phone'], request.form['commission_rate'] or 0,
          request.form['target_amount'] or 0, request.form['status'], request.form['hire_date'] or None,
          request.form['notes'], request.form['agent_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_sales_agents', request.form['agent_id'], None, {'agent_name': request.form['agent_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('sales_agents'))

@app.route('/sales-agents/delete/<int:agent_id>')
def delete_sales_agent(agent_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_sales_agents WHERE agent_id = %s", (agent_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_sales_agents', agent_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('sales_agents'))

@app.route('/sales-agents/update-achievement/<int:agent_id>', methods=['POST'])
def update_agent_achievement(agent_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_sales_agents SET achieved_amount = achieved_amount + %s WHERE agent_id = %s",
                   (request.form['amount'], agent_id))
    conn.commit()
    log_audit('UPDATE', 'tbl_sales_agents', agent_id, None, {'achieved_amount': request.form['amount']})
    cursor.close()
    conn.close()
    return redirect(url_for('sales_agents'))

# ============ PLANNER ============

@app.route('/planner')
def planner():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_planner ORDER BY event_date DESC")
    events = cursor.fetchall()
    cursor.execute("SELECT * FROM tbl_planner WHERE event_date >= CURDATE() AND event_date <= DATE_ADD(CURDATE(), INTERVAL 7 DAY) AND status = 'scheduled' ORDER BY event_date ASC, event_time ASC")
    upcoming_events = cursor.fetchall()
    cursor.execute("SELECT id, customer_name, proposed_work_date, jc_assigned_to FROM tbl_service_jc WHERE jc_closed IS NULL AND proposed_work_date IS NOT NULL ORDER BY proposed_work_date ASC")
    pending_jobs = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('planner.html', username=session['username'], role=session['user'],
                         events=events, upcoming_events=upcoming_events, pending_jobs=pending_jobs, active_page='planner')

@app.route('/planner/add', methods=['POST'])
def add_event():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO tbl_planner (event_title, event_description, event_date, event_time, end_date, event_type, related_jc_id, assigned_to, status, color, created_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (request.form['event_title'], request.form['event_description'], request.form['event_date'],
          request.form['event_time'] or None, request.form['end_date'] or None, request.form['event_type'],
          request.form['related_jc_id'] or None, request.form['assigned_to'] or None, request.form['status'],
          request.form['color'] or '#0066cc', session['username']))
    event_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_planner', event_id, None, {'event_title': request.form['event_title']})
    cursor.close()
    conn.close()
    return redirect(url_for('planner'))

@app.route('/planner/edit', methods=['POST'])
def edit_event():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE tbl_planner SET event_title = %s, event_description = %s, event_date = %s, event_time = %s,
        end_date = %s, event_type = %s, related_jc_id = %s, assigned_to = %s, status = %s, color = %s
        WHERE event_id = %s
    """, (request.form['event_title'], request.form['event_description'], request.form['event_date'],
          request.form['event_time'] or None, request.form['end_date'] or None, request.form['event_type'],
          request.form['related_jc_id'] or None, request.form['assigned_to'] or None, request.form['status'],
          request.form['color'] or '#0066cc', request.form['event_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_planner', request.form['event_id'], None, {'event_title': request.form['event_title']})
    cursor.close()
    conn.close()
    return redirect(url_for('planner'))

@app.route('/planner/delete/<int:event_id>')
def delete_event(event_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_planner WHERE event_id = %s", (event_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_planner', event_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('planner'))

@app.route('/planner/complete/<int:event_id>')
def complete_event(event_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE tbl_planner SET status = 'completed' WHERE event_id = %s", (event_id,))
    conn.commit()
    log_audit('UPDATE', 'tbl_planner', event_id, None, {'status': 'completed'})
    cursor.close()
    conn.close()
    return redirect(url_for('planner'))

# ============ EXPORT REPORTS ============

@app.route('/export/job-cards/excel')
def export_job_cards_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_service_jc ORDER BY id DESC")
    job_cards = cursor.fetchall()
    cursor.close()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Cards"
    ws.append(['ID', 'JC Number', 'Customer Name', 'Customer Type', 'JC Type', 'Amount', 'Assigned To', 'Status', 'Payment Status', 'Created Date'])
    for jc in job_cards:
        ws.append([jc['id'], jc.get('jc_number', ''), jc.get('customer_name', ''), jc.get('customer_type', ''),
                   jc.get('jc_type', ''), jc.get('amount', 0), jc.get('jc_assigned_to', 'Unassigned'),
                   'Closed' if jc.get('jc_closed') == 'Closed' else 'Open',
                   'Paid' if jc.get('paid') == 'Paid' else 'Pending',
                   jc.get('created_at').strftime('%Y-%m-%d') if jc.get('created_at') else ''])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='job_cards.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/job-cards/csv')
def export_job_cards_csv():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_service_jc ORDER BY id DESC")
    job_cards = cursor.fetchall()
    cursor.close()
    conn.close()
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'JC Number', 'Customer Name', 'Customer Type', 'JC Type', 'Amount', 'Assigned To', 'Status', 'Payment Status', 'Created Date'])
    for jc in job_cards:
        writer.writerow([jc['id'], jc.get('jc_number', ''), jc.get('customer_name', ''), jc.get('customer_type', ''),
                         jc.get('jc_type', ''), jc.get('amount', 0), jc.get('jc_assigned_to', 'Unassigned'),
                         'Closed' if jc.get('jc_closed') == 'Closed' else 'Open',
                         'Paid' if jc.get('paid') == 'Paid' else 'Pending',
                         jc.get('created_at').strftime('%Y-%m-%d') if jc.get('created_at') else ''])
    output.seek(0)
    return send_file(BytesIO(output.getvalue().encode('utf-8')), as_attachment=True, download_name='job_cards.csv', mimetype='text/csv')

@app.route('/export/customers/excel')
def export_customers_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_customer ORDER BY customer_id DESC")
    customers = cursor.fetchall()
    cursor.close()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    ws.append(['ID', 'Customer Name', 'Customer Type', 'Email', 'Phone', 'Created Date'])
    for c in customers:
        ws.append([c.get('customer_id'), c.get('customer_name', ''), c.get('customer_type', ''),
                   c.get('email', ''), c.get('phone', ''),
                   c.get('created_at').strftime('%Y-%m-%d') if c.get('created_at') else ''])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='customers.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/inventory/excel')
def export_inventory_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_item ORDER BY item_id DESC")
    items = cursor.fetchall()
    cursor.execute("SELECT * FROM tbl_product ORDER BY product_id DESC")
    products = cursor.fetchall()
    cursor.close()
    conn.close()
    wb = Workbook()
    ws_items = wb.active
    ws_items.title = "Items"
    ws_items.append(['ID', 'Item Name', 'Quantity', 'Purchase Price', 'Total Value'])
    for item in items:
        ws_items.append([item.get('item_id'), item.get('item_name', ''), item.get('quantity', 0),
                         item.get('purchase_price', 0), (item.get('quantity', 0) or 0) * (item.get('purchase_price', 0) or 0)])
    ws_products = wb.create_sheet("Products")
    ws_products.append(['ID', 'Product Name', 'Quantity', 'Purchase Price', 'Total Value'])
    for product in products:
        ws_products.append([product.get('product_id'), product.get('product_name', ''), product.get('quantity', 0),
                            product.get('purchase_price', 0), (product.get('quantity', 0) or 0) * (product.get('purchase_price', 0) or 0)])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='inventory.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ============ DAILY STATUS REPORT ============

@app.route('/daily-status')
def daily_status():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_daily_status ORDER BY report_date DESC LIMIT 30")
    reports = cursor.fetchall()
    today = date.today()
    cursor.execute("SELECT * FROM tbl_daily_status WHERE report_date = %s", (today,))
    today_report = cursor.fetchone()
    if not today_report:
        cursor.execute("""
            SELECT COUNT(*) as total_jobs, SUM(CASE WHEN DATE(created_at) = %s THEN 1 ELSE 0 END) as new_jobs,
                   SUM(CASE WHEN DATE(work_done_date) = %s AND jc_closed = 'Closed' THEN 1 ELSE 0 END) as closed_jobs,
                   SUM(CASE WHEN DATE(payment_date) = %s AND paid = 'Paid' THEN 1 ELSE 0 END) as paid_jobs,
                   COALESCE(SUM(CASE WHEN paid != 'Paid' AND jc_closed = 'Closed' THEN amount ELSE 0 END), 0) as pending_amount,
                   COALESCE(SUM(CASE WHEN paid = 'Paid' THEN total_paid_amount ELSE amount END), 0) as total_revenue
            FROM tbl_service_jc WHERE DATE(created_at) = %s OR DATE(work_done_date) = %s OR DATE(payment_date) = %s
        """, (today, today, today, today, today, today))
        metrics = cursor.fetchone()
        cursor.execute("""
            INSERT INTO tbl_daily_status (report_date, total_jobs, new_jobs, closed_jobs, paid_jobs, pending_payment_amount, total_revenue)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (today, metrics['total_jobs'] or 0, metrics['new_jobs'] or 0, metrics['closed_jobs'] or 0,
              metrics['paid_jobs'] or 0, metrics['pending_amount'] or 0, metrics['total_revenue'] or 0))
        conn.commit()
        cursor.execute("SELECT * FROM tbl_daily_status WHERE report_date = %s", (today,))
        today_report = cursor.fetchone()
    cursor.close()
    conn.close()
    return render_template('daily_status.html', username=session['username'], role=session['user'],
                         reports=reports, today_report=today_report, active_page='daily_status')

@app.route('/daily-status/generate')
def generate_daily_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    today = date.today()
    cursor.execute("""
        SELECT COUNT(*) as total_jobs, SUM(CASE WHEN DATE(created_at) = %s THEN 1 ELSE 0 END) as new_jobs,
               SUM(CASE WHEN DATE(work_done_date) = %s AND jc_closed = 'Closed' THEN 1 ELSE 0 END) as closed_jobs,
               SUM(CASE WHEN DATE(payment_date) = %s AND paid = 'Paid' THEN 1 ELSE 0 END) as paid_jobs,
               COALESCE(SUM(CASE WHEN paid != 'Paid' AND jc_closed = 'Closed' THEN amount ELSE 0 END), 0) as pending_amount,
               COALESCE(SUM(CASE WHEN paid = 'Paid' THEN total_paid_amount ELSE amount END), 0) as total_revenue
        FROM tbl_service_jc WHERE DATE(created_at) = %s OR DATE(work_done_date) = %s OR DATE(payment_date) = %s
    """, (today, today, today, today, today, today))
    metrics = cursor.fetchone()
    cursor.execute("""
        INSERT INTO tbl_daily_status (report_date, total_jobs, new_jobs, closed_jobs, paid_jobs, pending_payment_amount, total_revenue)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE total_jobs = VALUES(total_jobs), new_jobs = VALUES(new_jobs),
        closed_jobs = VALUES(closed_jobs), paid_jobs = VALUES(paid_jobs),
        pending_payment_amount = VALUES(pending_payment_amount), total_revenue = VALUES(total_revenue)
    """, (today, metrics['total_jobs'] or 0, metrics['new_jobs'] or 0, metrics['closed_jobs'] or 0,
          metrics['paid_jobs'] or 0, metrics['pending_amount'] or 0, metrics['total_revenue'] or 0))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('daily_status'))

@app.route('/daily-status/view/<report_date>')
def view_daily_report(report_date):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_daily_status WHERE report_date = %s", (report_date,))
    report = cursor.fetchone()
    cursor.execute("SELECT * FROM tbl_service_jc WHERE DATE(created_at) = %s OR DATE(work_done_date) = %s OR DATE(payment_date) = %s ORDER BY id DESC",
                   (report_date, report_date, report_date))
    jobs = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('view_daily_report.html', username=session['username'], role=session['user'],
                         report=report, jobs=jobs, report_date=report_date, active_page='daily_status')

# ============ PDF GENERATION ============

# Configure wkhtmltopdf path
# wkhtmltopdf_path = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
# config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path)
"""
@app.route('/pdf/job-card/<int:jc_id>')
def pdf_job_card(jc_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_service_jc WHERE id = %s", (jc_id,))
    job_card = cursor.fetchone()
    cursor.execute("SELECT * FROM tbl_mpesa_number")
    mpesa_numbers = cursor.fetchall()
    cursor.close()
    conn.close()
    rendered_html = render_template('pdf_job_card.html', job_card=job_card, mpesa_numbers=mpesa_numbers,
                                   username=session['username'], date=date.today())
    pdf_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdfkit.from_string(rendered_html, pdf_file.name, configuration=config)
    pdf_file.close()
    return send_file(pdf_file.name, as_attachment=True, download_name=f'job_card_{jc_id}.pdf')

@app.route('/pdf/financial-report')
def pdf_financial_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT COALESCE(SUM(CASE WHEN paid = 'Paid' THEN total_paid_amount ELSE amount END), 0) as total_revenue,
               COALESCE(SUM(CASE WHEN paid != 'Paid' AND jc_closed = 'Closed' THEN amount ELSE 0 END), 0) as pending_amount,
               COUNT(CASE WHEN paid = 'Paid' THEN 1 END) as paid_count, COUNT(*) as total_jobs
        FROM tbl_service_jc
    """)
    summary = cursor.fetchone()
    cursor.execute("SELECT * FROM tbl_service_jc ORDER BY created_at DESC LIMIT 50")
    payments = cursor.fetchall()
    cursor.close()
    conn.close()
    rendered_html = render_template('pdf_financial_report.html', summary=summary, payments=payments, date=date.today())
    pdf_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdfkit.from_string(rendered_html, pdf_file.name, configuration=config)
    pdf_file.close()
    return send_file(pdf_file.name, as_attachment=True, download_name='financial_report.pdf')

@app.route('/pdf/stock-report')
def pdf_stock_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_item ORDER BY item_name")
    items = cursor.fetchall()
    cursor.execute("SELECT * FROM tbl_product ORDER BY product_name")
    products = cursor.fetchall()
    cursor.execute("SELECT COALESCE(SUM(quantity * purchase_price), 0) as total FROM tbl_item")
    items_total = cursor.fetchone()
    cursor.execute("SELECT COALESCE(SUM(quantity * purchase_price), 0) as total FROM tbl_product")
    products_total = cursor.fetchone()
    cursor.close()
    conn.close()
    rendered_html = render_template('pdf_stock_report.html', items=items, products=products,
                                   items_total=items_total['total'], products_total=products_total['total'], date=date.today())
    pdf_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdfkit.from_string(rendered_html, pdf_file.name, configuration=config)
    pdf_file.close()
    return send_file(pdf_file.name, as_attachment=True, download_name='stock_report.pdf')
"""
# ============ AUDIT LOG ============

@app.route('/audit-log')
def audit_log():
    if 'user_id' not in session or session.get('user') != 'admin':
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    action = request.args.get('action', '')
    table = request.args.get('table', '')
    username = request.args.get('username', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    query = "SELECT * FROM tbl_audit_log WHERE 1=1"
    params = []
    if action:
        query += " AND action = %s"
        params.append(action)
    if table:
        query += " AND table_name = %s"
        params.append(table)
    if username:
        query += " AND username LIKE %s"
        params.append(f"%{username}%")
    if date_from:
        query += " AND DATE(created_at) >= %s"
        params.append(date_from)
    if date_to:
        query += " AND DATE(created_at) <= %s"
        params.append(date_to)
    query += " ORDER BY created_at DESC LIMIT 500"
    cursor.execute(query, params)
    logs = cursor.fetchall()
    cursor.execute("SELECT DISTINCT action FROM tbl_audit_log")
    actions = [row['action'] for row in cursor.fetchall()]
    cursor.execute("SELECT DISTINCT table_name FROM tbl_audit_log")
    tables = [row['table_name'] for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    return render_template('audit_log.html', username=session['username'], role=session['user'],
                         logs=logs, actions=actions, tables=tables, active_page='audit_log')


# ============ SUPPLIERS ============

@app.route('/suppliers')
def suppliers():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_supplier ORDER BY supplier_id DESC")
    suppliers = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('suppliers.html', username=session['username'], role=session['user'],
                         suppliers=suppliers, active_page='suppliers')

@app.route('/suppliers/add', methods=['POST'])
def add_supplier():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO tbl_supplier (supplier_name, contact_person, email, phone, address, payment_terms, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (request.form['supplier_name'], request.form['contact_person'], request.form['email'],
          request.form['phone'], request.form['address'], request.form['payment_terms'], request.form['status']))
    supplier_id = cursor.lastrowid
    conn.commit()
    log_audit('CREATE', 'tbl_supplier', supplier_id, None, {'supplier_name': request.form['supplier_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('suppliers'))

@app.route('/suppliers/edit', methods=['POST'])
def edit_supplier():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE tbl_supplier SET supplier_name = %s, contact_person = %s, email = %s, phone = %s,
        address = %s, payment_terms = %s, status = %s WHERE supplier_id = %s
    """, (request.form['supplier_name'], request.form['contact_person'], request.form['email'],
          request.form['phone'], request.form['address'], request.form['payment_terms'],
          request.form['status'], request.form['supplier_id']))
    conn.commit()
    log_audit('UPDATE', 'tbl_supplier', request.form['supplier_id'], None, {'supplier_name': request.form['supplier_name']})
    cursor.close()
    conn.close()
    return redirect(url_for('suppliers'))

@app.route('/suppliers/delete/<int:supplier_id>')
def delete_supplier(supplier_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_supplier WHERE supplier_id = %s", (supplier_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_supplier', supplier_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('suppliers'))

# ============ PURCHASE ORDERS ============

@app.route('/purchase-orders')
def purchase_orders():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_purchase_order ORDER BY po_id DESC")
    purchase_orders = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('purchase_orders.html', username=session['username'], role=session['user'],
                         purchase_orders=purchase_orders, active_page='purchase_orders')

@app.route('/purchase-orders/create', methods=['GET'])
def create_po_form():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT supplier_id, supplier_name FROM tbl_supplier WHERE status = 'active'")
    suppliers = cursor.fetchall()
    cursor.execute("SELECT item_id, item_name, quantity, purchase_price FROM tbl_item")
    items = cursor.fetchall()
    cursor.execute("SELECT product_id, product_name, quantity, purchase_price FROM tbl_product")
    products = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('create_po.html', username=session['username'], role=session['user'],
                         suppliers=suppliers, items=items, products=products, now=date.today(),
                         active_page='purchase_orders')

@app.route('/purchase-orders/create', methods=['POST'])
def create_po_post():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    po_number = f"PO-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    cursor.execute("""
        INSERT INTO tbl_purchase_order (po_number, supplier_id, supplier_name, order_date, expected_delivery, notes, created_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (po_number, request.form['supplier_id'], request.form['supplier_name'],
          request.form['order_date'], request.form['expected_delivery'], request.form['notes'], session['username']))
    po_id = cursor.lastrowid
    
    # Add items
    item_ids = request.form.getlist('item_id')
    item_quantities = request.form.getlist('item_quantity')
    item_prices = request.form.getlist('item_price')
    for i in range(len(item_ids)):
        if item_ids[i] and item_quantities[i]:
            total = float(item_quantities[i]) * float(item_prices[i])
            cursor.execute("""
                INSERT INTO tbl_purchase_order_item (po_id, item_type, item_id, item_name, quantity, unit_price, total_price)
                VALUES (%s, 'item', %s, %s, %s, %s, %s)
            """, (po_id, item_ids[i], '', item_quantities[i], item_prices[i], total))
    
    # Add products
    product_ids = request.form.getlist('product_id')
    product_quantities = request.form.getlist('product_quantity')
    product_prices = request.form.getlist('product_price')
    for i in range(len(product_ids)):
        if product_ids[i] and product_quantities[i]:
            total = float(product_quantities[i]) * float(product_prices[i])
            cursor.execute("""
                INSERT INTO tbl_purchase_order_item (po_id, item_type, item_id, item_name, quantity, unit_price, total_price)
                VALUES (%s, 'product', %s, %s, %s, %s, %s)
            """, (po_id, product_ids[i], '', product_quantities[i], product_prices[i], total))
    
    conn.commit()
    log_audit('CREATE', 'tbl_purchase_order', po_id, None, {'po_number': po_number})
    cursor.close()
    conn.close()
    return redirect(url_for('purchase_orders'))

@app.route('/purchase-orders/receive/<int:po_id>', methods=['GET'])
def receive_material_form(po_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_purchase_order WHERE po_id = %s", (po_id,))
    po = cursor.fetchone()
    cursor.execute("SELECT * FROM tbl_purchase_order_item WHERE po_id = %s", (po_id,))
    items = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('receive_material.html', username=session['username'], role=session['user'],
                         po=po, items=items, active_page='purchase_orders')

@app.route('/purchase-orders/receive', methods=['POST'])
def receive_material_post():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    po_id = request.form['po_id']
    po_number = request.form['po_number']
    supplier_id = request.form['supplier_id']
    supplier_name = request.form['supplier_name']
    receive_date = request.form['receive_date']
    
    item_ids = request.form.getlist('item_id')
    item_types = request.form.getlist('item_type')
    item_names = request.form.getlist('item_name')
    quantities = request.form.getlist('quantity')
    prices = request.form.getlist('price')
    
    for i in range(len(item_ids)):
        if item_ids[i] and quantities[i] and int(quantities[i]) > 0:
            total = float(quantities[i]) * float(prices[i])
            cursor.execute("""
                INSERT INTO tbl_material_received (po_id, po_number, supplier_id, supplier_name, receive_date,
                item_type, item_id, item_name, quantity_received, unit_price, total_price, received_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (po_id, po_number, supplier_id, supplier_name, receive_date, item_types[i],
                  item_ids[i], item_names[i], quantities[i], prices[i], total, session['username']))
            
            # Update stock
            if item_types[i] == 'item':
                cursor.execute("UPDATE tbl_item SET quantity = quantity + %s WHERE item_id = %s",
                              (quantities[i], item_ids[i]))
            else:
                cursor.execute("UPDATE tbl_product SET quantity = quantity + %s WHERE product_id = %s",
                              (quantities[i], item_ids[i]))
            
            # Update PO item received quantity
            cursor.execute("""
                UPDATE tbl_purchase_order_item SET received_quantity = received_quantity + %s
                WHERE po_id = %s AND item_id = %s AND item_type = %s
            """, (quantities[i], po_id, item_ids[i], item_types[i]))
    
    cursor.execute("UPDATE tbl_purchase_order SET status = 'received' WHERE po_id = %s", (po_id,))
    conn.commit()
    log_audit('UPDATE', 'tbl_purchase_order', po_id, None, {'status': 'received'})
    cursor.close()
    conn.close()
    return redirect(url_for('purchase_orders'))

@app.route('/purchase-orders/delete/<int:po_id>')
def delete_po(po_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_purchase_order_item WHERE po_id = %s", (po_id,))
    cursor.execute("DELETE FROM tbl_purchase_order WHERE po_id = %s", (po_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_purchase_order', po_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('purchase_orders'))


# ============ QUOTATIONS MODULE ============

@app.route('/quotations')
def quotations():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_quotation ORDER BY quote_id DESC")
    quotations = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('quotations.html', username=session['username'], role=session['user'],
                         quotations=quotations, active_page='quotations')

@app.route('/quotations/create', methods=['GET'])
def create_quotation_form():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT customer_id, customer_name, email, phone FROM tbl_customer WHERE customer_type != 'Reseller'")
    customers = cursor.fetchall()
    cursor.execute("SELECT item_id, item_name, quantity, purchase_price FROM tbl_item WHERE quantity > 0")
    items = cursor.fetchall()
    cursor.execute("SELECT product_id, product_name, quantity, purchase_price FROM tbl_product WHERE quantity > 0")
    products = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('create_quotation.html', username=session['username'], role=session['user'],
                         customers=customers, items=items, products=products, now=date.today(),
                         active_page='quotations')

@app.route('/quotations/view/<int:quote_id>')
def view_quotation(quote_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_quotation WHERE quote_id = %s", (quote_id,))
    quote = cursor.fetchone()
    cursor.execute("SELECT * FROM tbl_quotation_item WHERE quote_id = %s", (quote_id,))
    items = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('view_quotation.html', username=session['username'], role=session['user'],
                         quote=quote, items=items, active_page='quotations')

@app.route('/quotations/edit/<int:quote_id>', methods=['GET'])
def edit_quotation_form(quote_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_quotation WHERE quote_id = %s", (quote_id,))
    quote = cursor.fetchone()
    cursor.execute("SELECT * FROM tbl_quotation_item WHERE quote_id = %s", (quote_id,))
    items = cursor.fetchall()
    cursor.execute("SELECT customer_id, customer_name, email, phone FROM tbl_customer")
    customers = cursor.fetchall()
    cursor.execute("SELECT item_id, item_name, quantity, purchase_price FROM tbl_item")
    all_items = cursor.fetchall()
    cursor.execute("SELECT product_id, product_name, quantity, purchase_price FROM tbl_product")
    all_products = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('edit_quotation.html', username=session['username'], role=session['user'],
                         quote=quote, items=items, customers=customers, all_items=all_items,
                         all_products=all_products, active_page='quotations')

@app.route('/quotations/edit/<int:quote_id>', methods=['POST'])
def edit_quotation_post(quote_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        UPDATE tbl_quotation SET customer_id = %s, customer_name = %s, customer_email = %s, customer_phone = %s,
        quote_date = %s, expiry_date = %s, subtotal = %s, tax_rate = %s, tax_amount = %s,
        discount_amount = %s, total_amount = %s, notes = %s, terms_conditions = %s, status = %s
        WHERE quote_id = %s
    """, (request.form['customer_id'] or None, request.form['customer_name'], request.form['customer_email'],
          request.form['customer_phone'], request.form['quote_date'], request.form['expiry_date'] or None,
          request.form['subtotal'] or 0, request.form['tax_rate'] or 0, request.form['tax_amount'] or 0,
          request.form['discount_amount'] or 0, request.form['total_amount'] or 0,
          request.form['notes'], request.form['terms_conditions'], request.form['status'], quote_id))
    
    # Delete old items and insert new ones
    cursor.execute("DELETE FROM tbl_quotation_item WHERE quote_id = %s", (quote_id,))
    
    item_ids = request.form.getlist('item_id')
    item_quantities = request.form.getlist('item_quantity')
    item_prices = request.form.getlist('item_price')
    
    for i in range(len(item_ids)):
        if item_ids[i] and item_quantities[i] and int(item_quantities[i]) > 0:
            total = float(item_quantities[i]) * float(item_prices[i])
            cursor.execute("""
                INSERT INTO tbl_quotation_item (quote_id, item_type, item_id, item_name, quantity, unit_price, total_price)
                VALUES (%s, 'item', %s, %s, %s, %s, %s)
            """, (quote_id, item_ids[i], '', item_quantities[i], item_prices[i], total))
    
    product_ids = request.form.getlist('product_id')
    product_quantities = request.form.getlist('product_quantity')
    product_prices = request.form.getlist('product_price')
    
    for i in range(len(product_ids)):
        if product_ids[i] and product_quantities[i] and int(product_quantities[i]) > 0:
            total = float(product_quantities[i]) * float(product_prices[i])
            cursor.execute("""
                INSERT INTO tbl_quotation_item (quote_id, item_type, item_id, item_name, quantity, unit_price, total_price)
                VALUES (%s, 'product', %s, %s, %s, %s, %s)
            """, (quote_id, product_ids[i], '', product_quantities[i], product_prices[i], total))
    
    conn.commit()
    log_audit('UPDATE', 'tbl_quotation', quote_id, None, {'status': request.form['status']})
    cursor.close()
    conn.close()
    
    return redirect(url_for('quotations'))

@app.route('/quotations/update-status/<int:quote_id>', methods=['POST'])
def update_quotation_status(quote_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    new_status = request.form['status']
    cursor.execute("UPDATE tbl_quotation SET status = %s WHERE quote_id = %s", (new_status, quote_id))
    conn.commit()
    log_audit('UPDATE', 'tbl_quotation', quote_id, None, {'status': new_status})
    cursor.close()
    conn.close()
    return redirect(url_for('view_quotation', quote_id=quote_id))

@app.route('/quotations/convert/<int:quote_id>')
def convert_quote_to_jc(quote_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    # Get quotation
    cursor.execute("SELECT * FROM tbl_quotation WHERE quote_id = %s", (quote_id,))
    quote = cursor.fetchone()
    
    # Create job card from quotation
    cursor.execute("""
        INSERT INTO tbl_service_jc (jc_type, customer_name, customer_id, amount, work_statement, created_by)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, ('Service', quote['customer_name'], quote['customer_id'], quote['total_amount'], 
          f"Converted from Quotation #{quote['quote_number']}", session['username']))
    
    jc_id = cursor.lastrowid
    
    # Get quotation items
    cursor.execute("SELECT * FROM tbl_quotation_item WHERE quote_id = %s", (quote_id,))
    items = cursor.fetchall()
    
    for item in items:
        cursor.execute("""
            INSERT INTO tbl_service_jc_item (service_jc_id, item_id, product_id, item_name, product_name, item_quantity, product_quantity)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (jc_id, item['item_id'] if item['item_type'] == 'item' else None,
              item['item_id'] if item['item_type'] == 'product' else None,
              item['item_name'], item['item_name'], item['quantity'], item['quantity']))
    
    # Update quotation as converted
    cursor.execute("UPDATE tbl_quotation SET status = 'converted', converted_jc_id = %s WHERE quote_id = %s",
                  (jc_id, quote_id))
    
    conn.commit()
    log_audit('CONVERT', 'tbl_quotation', quote_id, None, {'converted_jc_id': jc_id})
    cursor.close()
    conn.close()
    
    return redirect(url_for('view_job_card', jc_id=jc_id))

@app.route('/quotations/delete/<int:quote_id>')
def delete_quotation(quote_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM tbl_quotation_item WHERE quote_id = %s", (quote_id,))
    cursor.execute("DELETE FROM tbl_quotation WHERE quote_id = %s", (quote_id,))
    conn.commit()
    log_audit('DELETE', 'tbl_quotation', quote_id, None, None)
    cursor.close()
    conn.close()
    return redirect(url_for('quotations'))

@app.route('/pdf/quotation/<int:quote_id>')
def pdf_quotation(quote_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_quotation WHERE quote_id = %s", (quote_id,))
    quote = cursor.fetchone()
    cursor.execute("SELECT * FROM tbl_quotation_item WHERE quote_id = %s", (quote_id,))
    items = cursor.fetchall()
    cursor.close()
    conn.close()
    
    rendered_html = render_template('pdf_quotation.html', quote=quote, items=items, date=date.today())
    pdf_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdfkit.from_string(rendered_html, pdf_file.name, configuration=config)
    pdf_file.close()
    return send_file(pdf_file.name, as_attachment=True, download_name=f'quotation_{quote["quote_number"]}.pdf')


@app.route('/quotations/create', methods=['POST'])
def create_quotation_post():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    cursor = conn.cursor()
    
    quote_number = f"QT-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # Get form data
    subtotal = float(request.form['subtotal'] or 0)
    tax_rate = float(request.form['tax_rate'] or 0)
    tax_amount = float(request.form['tax_amount'] or 0)
    discount_amount = float(request.form['discount_amount'] or 0)
    total_amount = float(request.form['total_amount'] or 0)
    
    # Calculate cost price (sum of purchase prices of items)
    item_ids = request.form.getlist('item_id')
    item_quantities = request.form.getlist('item_quantity')
    product_ids = request.form.getlist('product_id')
    product_quantities = request.form.getlist('product_quantity')
    
    cost_price = 0
    
    # Calculate item costs
    for i in range(len(item_ids)):
        if item_ids[i] and item_quantities[i] and int(item_quantities[i]) > 0:
            cursor.execute("SELECT purchase_price FROM tbl_item WHERE item_id = %s", (item_ids[i],))
            item = cursor.fetchone()
            if item:
                cost_price += float(item[0]) * int(item_quantities[i])
    
    # Calculate product costs
    for i in range(len(product_ids)):
        if product_ids[i] and product_quantities[i] and int(product_quantities[i]) > 0:
            cursor.execute("SELECT purchase_price FROM tbl_product WHERE product_id = %s", (product_ids[i],))
            product = cursor.fetchone()
            if product:
                cost_price += float(product[0]) * int(product_quantities[i])
    
    # Calculate profit and margin
    gross_profit = total_amount - cost_price
    margin_percentage = (gross_profit / total_amount * 100) if total_amount > 0 else 0
    
    cursor.execute("""
        INSERT INTO tbl_quotation (quote_number, customer_id, customer_name, customer_email, customer_phone,
        quote_date, expiry_date, subtotal, cost_price, gross_profit, margin_percentage,
        tax_rate, tax_amount, discount_amount, total_amount, notes, terms_conditions, status, created_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (quote_number, request.form['customer_id'] or None, request.form['customer_name'],
          request.form['customer_email'], request.form['customer_phone'], request.form['quote_date'],
          request.form['expiry_date'] or None, subtotal, cost_price, gross_profit, margin_percentage,
          tax_rate, tax_amount, discount_amount, total_amount, request.form['notes'],
          request.form['terms_conditions'], 'draft', session['username']))
    
    quote_id = cursor.lastrowid
    
    # Add items (same as before)
    for i in range(len(item_ids)):
        if item_ids[i] and item_quantities[i] and int(item_quantities[i]) > 0:
            cursor.execute("SELECT item_name, purchase_price FROM tbl_item WHERE item_id = %s", (item_ids[i],))
            item = cursor.fetchone()
            total = float(item_quantities[i]) * float(request.form.getlist('item_price')[i])
            cursor.execute("""
                INSERT INTO tbl_quotation_item (quote_id, item_type, item_id, item_name, quantity, unit_price, total_price)
                VALUES (%s, 'item', %s, %s, %s, %s, %s)
            """, (quote_id, item_ids[i], item[0], item_quantities[i], request.form.getlist('item_price')[i], total))
    
    for i in range(len(product_ids)):
        if product_ids[i] and product_quantities[i] and int(product_quantities[i]) > 0:
            cursor.execute("SELECT product_name, purchase_price FROM tbl_product WHERE product_id = %s", (product_ids[i],))
            product = cursor.fetchone()
            total = float(product_quantities[i]) * float(request.form.getlist('product_price')[i])
            cursor.execute("""
                INSERT INTO tbl_quotation_item (quote_id, item_type, item_id, item_name, quantity, unit_price, total_price)
                VALUES (%s, 'product', %s, %s, %s, %s, %s)
            """, (quote_id, product_ids[i], product[0], product_quantities[i], request.form.getlist('product_price')[i], total))
    
    conn.commit()
    log_audit('CREATE', 'tbl_quotation', quote_id, None, {'quote_number': quote_number})
    cursor.close()
    conn.close()
    
    return redirect(url_for('quotations'))

import os

# For Render.com compatibility
port = int(os.environ.get('PORT', 5000))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=port, debug=False)